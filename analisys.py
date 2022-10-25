from abc import ABC, abstractmethod
import os
import openpyxl
from settings import Const


class NoAbsolutePath(Exception):
    pass


class Specification(ABC):

    @abstractmethod
    def is_satisfied(self, parameter):
        pass


class CustomFilter(ABC):

    @abstractmethod
    def custom_filter(self, articles, specification):
        pass


class Book(openpyxl.workbook.workbook.Workbook):

    EXTENSIONS = ('.xlsx',
                  '.xls')
    data_sheet = None

    def __init__(self, path_to_file):
        super().__init__()
        self.path = path_to_file
        self.book = openpyxl.Workbook()
        self.active_sheet = None

    @classmethod
    def is_correct_path(cls, path: str):
        """ Create an object of class Book if path is correct.

        :param path: the absolute path to a xlsx-file
        :type path: string

        """
        if os.path.isabs(path):
            if os.path.isfile(path) and os.path.splitext(path)[1] in cls.EXTENSIONS:
                return cls(path)
            else:
                raise FileExistsError
        else:
            raise NoAbsolutePath

    def open_book(self):
        self.book = openpyxl.load_workbook(self.path)

    def assign_data_sheet(self, sheet_name):
        self.data_sheet = self.book[sheet_name]

    def make_active_sheet(self, sheet_name: str):

        """ make active sheet = 'sheet_name' or if it doesn't exist crete it """

        if sheet_name not in self.book.sheetnames:
            self.book.create_sheet(title=sheet_name)
        self.active_sheet = self.book[sheet_name]
        self.save_book()

    def save_book(self):
        self.book.save(self.path)


class MeasureSpec(Specification):
    def __init__(self, condition):
        self.condition = condition

    def is_satisfied(self, article):
        return article.parameter_list[3] == self.condition


class AndSpecification(Specification):
    def __init__(self, *args):
        self.args = args

    def is_satisfied(self, parameter):
        return all(map(
            lambda specification: specification.is_satisfied(parameter), self.args
        ))


class AnySpecification(Specification):
    def __init__(self, *args):
        self.args = args

    def is_satisfied(self, parameter):
        return any(map(
            lambda specification: specification.is_satisfied(parameter), self.args
        ))


class WorkingFilter(CustomFilter):
    def custom_filter(self, articles, specification):
        for article in articles:
            if specification.is_satisfied(article):
                yield article


class Article:
    def __init__(self, parameter_list):
        self.parameter_list = parameter_list


class Writer(ABC):

    @abstractmethod
    def update_book(self):
        pass

    @abstractmethod
    def create_sheet(self, sheet=Const.WRITER_SETTINGS['new_sheet_name']):
        pass

    @abstractmethod
    def write_to(self):
        pass


class SimpleWriter(Writer):

    def __init__(self, articles, **kwargs):
        self.articles = articles
        self.kwargs = kwargs
        self.book = None
        if self.kwargs['new_book_name'] is not None:
            self.book_path = fr'data\{kwargs["new_book_name"]}.xlsx'
            self.create_book()
        else:
            self.book_path = Const.BOOK_PATH

    def create_book(self):
        book = Book(self.book_path)
        book.save_book()
        self.book = book
        self.book.close()

    def update_book(self):
        book = Book.is_correct_path(os.path.abspath(self.book_path))
        self.book = book
        book.open_book()

    def create_sheet(self, sheet=Const.WRITER_SETTINGS['new_sheet_name']):
        self.book.make_active_sheet(sheet)

    def write_to(self):
        def sequence(articles):
            for item in articles:
                yield item.parameter_list

        articles = sequence(self.articles)
        #article = next(articles)
        for row in range(1, (len(self.articles) * 2)):
            if row % 2 == 0:  #: Second_row
                for column in range(self.kwargs['column_start'], (len(list(article.items())[1][1]))):
                    value = list(article.items())[1][1][column - 1]
                    if value is None:
                        value = ''
                    cell = self.book.active_sheet.cell(row=row + self.kwargs['row_start'] - 1, column=column)
                    cell.value = value
            else:  #: First_row
                article = next(articles)
                for column in range(self.kwargs['column_start'], (len(list(article.items())[1][1]))):
                    value = list(article.items())[0][1][column - 1]
                    if value is None:
                        value = ''
                    cell = self.book.active_sheet.cell(row=row + self.kwargs['row_start'] - 1, column=column)
                    cell.value = value


class PipesDiameterSpecification(Specification):
    def __init__(self, condition):
        self.condition = condition

    def is_satisfied(self, article):
        param = article.parameter_list[2].split(':')
        value = param[0].split(',')[4].strip() + ':' + param[1]
        if self.condition in value:
            return True
